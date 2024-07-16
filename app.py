from flask import Flask, render_template, request

import openpyxl
import magic
import io  # Used for handling file-like objects

# Improved error handling with custom exception classes
class InvalidFileTypeError(Exception):
    pass

class FileValidationError(Exception):
    def __init__(self, errors):
        self.errors = errors

app = Flask(__name__, template_folder='templates')  # Specify template folder

@app.route('/')
def index():
    return render_template('index.html')  # Ensure correct file name

@app.route('/analyze_excel', methods=['POST'])
def analyze_excel():
    if 'file' not in request.files:
        return render_template('index.html', error='No file uploaded'), 400

    file = request.files['file']

    # Check file type using filename extension (optional)
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        return render_template('index.html', error='Only Excel files (.xls or .xlsx) are allowed'), 400

    # Convert uploaded file to a file-like object
    file_data = io.BytesIO(file.read())
    file_data.seek(0)  # Reset the pointer to the beginning

    try:
        workbook = openpyxl.load_workbook(file_data, data_only=True)
        worksheet = workbook.active
    except Exception as e:  # Catch any exception related to file opening
        return render_template('index.html', error=f'An error occurred while processing the file: {str(e)}'), 400

    # Assume the header row exists on the first row (adjust as needed)
    header_row = [cell.value.strip() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    mobile_number_index = header_row.index('Mobile Number')

    # Function to validate a single row
    def validate_row(row):
        errors = []
        # Field validations based on your requirements
        if not row[header_row.index('First Name')].value.strip():
            errors.append({'field': 'First Name', 'message': 'Missing mandatory field'})
        else:
            if isinstance(row[header_row.index('First Name')].value, str):
            # Only call strip() on strings
                if not row[header_row.index('First Name')].value.strip():
                    errors.append({'field': 'First Name', 'message': 'Missing mandatory field'})
            else:
            # Handle case where expected value is a string but cell contains a number
                errors.append({'field': 'First Name', 'message': 'Invalid format: Must be text'})

        if not row[header_row.index('Country Code')].value:  # Check for empty value (including whitespace)
            errors.append({'field': 'Country Code', 'message': 'Missing mandatory field'})
        else:
         # You can add validation for Country Code format here (e.g., length)
         country_code = str(row[header_row.index('Country Code')].value)  # Convert to string
        if len(country_code) != 2:
            errors.append({'field': 'Country Code', 'message': 'Invalid format: Must be 2 characters'})

            # **Fix applied here:**
            # Ensure conversion to string before stripping for Mobile Number
            if row[mobile_number_index].value is None:
                errors.append({'field': 'Mobile Number', 'message': 'Missing mandatory field'})
            else:
                mobile_number = str(row[mobile_number_index].value)  # Convert to string
                mobile_number = mobile_number.strip()
                if not isinstance(mobile_number, str):
                    errors.append({'field': 'Mobile Number', 'message': 'Invalid format: Must be text'})

        if not row[header_row.index('Address')].value:  # Check for empty value (including whitespace)
            errors.append({'field': 'Address', 'message': 'Missing mandatory field'})
        # No further validation needed for address (assuming any format is allowed)

        # Last Name and Notes are optional, so no validation required

        if errors:
            raise FileValidationError(errors)

        country_code = str(row[header_row.index('Country Code')].value)  # Convert to string
        mobile_number = str(row[mobile_number_index].value)  # Convert to string (fixed line)

        return {
            'data': {
                'First Name': row[header_row.index('First Name')].value.strip(),
                'Last Name': row[header_row.index('Last Name')].value.strip() if row[header_row.index('Last Name')] else None,
                'Country Code' : country_code.strip(),  # Now call strip() on the string
                'Mobile Number': row[header_row.index('Mobile Number')].value,
                'Address': row[header_row.index('Address')].value.strip(),
                'Notes': row[header_row.index('Notes')].value.strip() if row[header_row.index('Notes')] else None,
            }
        }

    data = {}
    for row in worksheet.iter_rows(min_row=2):  # Skip header row (adjust as needed)
        try:
            data = validate_row(row)
            print(data)  # Print data for debugging
        except FileValidationError as e:
            # Handle validation errors here (e.g., log errors, display in UI)
            pass  # Simulate handling for now

    return render_template('index.html', data=data)
    